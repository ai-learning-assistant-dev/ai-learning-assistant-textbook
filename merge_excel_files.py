import argparse
import shutil
import os
import openpyxl
import re

def get_video_key(url):
    """
    从URL中提取视频唯一标识 (BVID, p)
    使用正则匹配，比urlparse更健壮
    """
    if not url or not isinstance(url, str):
        return None
    
    try:
        # 1. 提取BVID (如 BV1jF4SzDEJ5)
        # BVID通常是BV开头，后面跟10位左右字符
        bvid_match = re.search(r'(BV[a-zA-Z0-9]+)', url, re.IGNORECASE)
        if not bvid_match:
            return None
        bvid = bvid_match.group(1).upper() # 统一大写
        
        # 2. 提取p参数
        # 常见格式: ?p=2, &p=2, /?p=2
        # 如果没有p参数，默认为'1'
        p = '1'
        p_match = re.search(r'[?&]p=(\d+)', url)
        if p_match:
            p = p_match.group(1)
            
        return (bvid, p)
    except Exception as e:
        # print(f"URL解析错误: {url} -> {e}")
        return None

def merge_excel_files(manual_file, program_file, output_file=None):
    # 1. 确定输出文件路径
    if not output_file:
        output_file = "merged_output.xlsx"
    
    print("="*60)
    print(f"开始合并任务")
    print(f"人工Excel: {manual_file}")
    print(f"程序Excel: {program_file}")
    print(f"输出文件: {output_file}")
    print("="*60)
    
    # 2. 初始化输出文件：复制人工Excel
    try:
        shutil.copy(manual_file, output_file)
        print(f"✅ 已复制人工Excel到输出路径")
    except Exception as e:
        print(f"❌ 错误: 无法复制人工Excel文件: {e}")
        return

    try:
        # 加载工作簿
        print(f"正在加载Excel文件...")
        wb_out = openpyxl.load_workbook(output_file)
        wb_prog = openpyxl.load_workbook(program_file)
        
        # ==========================================
        # 处理 chapters_sections 分表
        # ==========================================
        if 'chapters_sections' in wb_out.sheetnames and 'chapters_sections' in wb_prog.sheetnames:
            print("\n正在处理 chapters_sections 分表...")
            ws_out = wb_out['chapters_sections']
            ws_prog = wb_prog['chapters_sections']
            
            # 1. 构建程序Excel的数据索引
            prog_data_map = {}
            prog_rows_count = 0
            
            # 获取程序表的列索引映射 (假设第一行是表头)
            # 我们需要: 视频URL, 节标题, 预设问题1, 预设问题2, 预设问题3
            # 虽然用户描述了固定顺序，但动态查找更安全，如果没找到则回退到固定索引
            prog_header = [c.value for c in ws_prog[1]]
            
            # 辅助函数：获取列索引（0-based），如果没找到返回默认值
            def get_col_idx(headers, names, default):
                for i, h in enumerate(headers):
                    if h and str(h).strip() in names:
                        return i
                return default

            # 定义列索引 (0-based for list access, but openpyxl values_only gives tuple)
            # 程序表结构: "序号 章标题 章顺序 视频URL 节标题 节顺序 课时/min 预设问题1 预设问题2 预设问题3 知识内容"
            # 对应的索引: 3=URL, 4=节标题, 7=Q1, 8=Q2, 9=Q3
            p_idx_url = get_col_idx(prog_header, ['视频URL', 'URL'], 3)
            p_idx_title = get_col_idx(prog_header, ['节标题'], 4)
            p_idx_q1 = get_col_idx(prog_header, ['预设问题1'], 7)
            p_idx_q2 = get_col_idx(prog_header, ['预设问题2'], 8)
            p_idx_q3 = get_col_idx(prog_header, ['预设问题3'], 9)
            
            print(f"程序表列索引检测: URL={p_idx_url}, 节标题={p_idx_title}, Q1={p_idx_q1}")

            for row in ws_prog.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(p_idx_url, p_idx_title): 
                    continue
                
                url = row[p_idx_url]
                key = get_video_key(url)
                
                if key:
                    prog_rows_count += 1
                    q1 = row[p_idx_q1] if len(row) > p_idx_q1 else None
                    q2 = row[p_idx_q2] if len(row) > p_idx_q2 else None
                    q3 = row[p_idx_q3] if len(row) > p_idx_q3 else None
                    section_title = row[p_idx_title]
                    
                    # 存入字典
                    prog_data_map[key] = {
                        'questions': [q1, q2, q3],
                        'section_title': section_title,
                        'raw_url': url # debug用
                    }
            
            print(f"已索引程序表数据: {len(prog_data_map)} 条有效记录")

            # 2. 准备输出表（人工表）
            # 确定列索引
            out_header_cells = ws_out[1]
            out_header = [c.value for c in out_header_cells]
            
            # 查找 视频URL 列
            o_idx_url_col_num = 4 # Default Column 4
            for cell in out_header_cells:
                if cell.value and str(cell.value).strip() in ['视频URL', 'URL']:
                    o_idx_url_col_num = cell.column
                    break
            
            # 查找 预设问题 列 (我们需要列号 1-based)
            o_col_q1 = 8
            o_col_q2 = 9
            o_col_q3 = 10
            
            # 查找/创建 机械标题 列
            mech_title_col_idx = 12 # 默认第12列
            found_mech_col = False
            
            # 动态计算最大列数，确保追加在最后
            max_col = ws_out.max_column
            
            # 遍历表头寻找是否已存在
            for cell in out_header_cells:
                if cell.value and str(cell.value).strip() == "机械标题":
                    mech_title_col_idx = cell.column
                    found_mech_col = True
                    print(f"✅ 找到现有'机械标题'列: 第 {mech_title_col_idx} 列")
                    break
            
            if not found_mech_col:
                # 如果没找到，则追加在现有最后一列的后面
                # 注意：如果max_col < 11，直接用max_col+1可能会导致数据挤在一起
                # 但为了保证用户能看见，追加在最后是最稳妥的
                # 同时保留至少12列的结构（如果原表很短）
                mech_title_col_idx = max(max_col + 1, 12)
                
                print(f"⚠️ 未找到'机械标题'列，将在第 {mech_title_col_idx} 列创建 (当前最大列: {max_col})")
                # 强制写入表头
                ws_out.cell(row=1, column=mech_title_col_idx, value="机械标题")
            
            # 关键修复：手动更新 ws_out.max_column，防止后续迭代无法覆盖新列
            # 虽然 openpyxl 理论上会自动更新，但在某些模式下可能不会立即生效
            # 我们通过 mech_title_col_idx 来确保这一列是有效的
            
            # 3. 遍历并更新
            matched_count = 0
            processed_count = 0
            
            for row in ws_out.iter_rows(min_row=2):
                processed_count += 1
                # 获取URL单元格
                # 注意：iter_rows返回的是单元格对象元组
                if len(row) < o_idx_url_col_num:
                    continue
                    
                url_cell = row[o_idx_url_col_num - 1] # 0-based index
                url = url_cell.value
                
                if not url:
                    continue
                
                key = get_video_key(url)
                current_row_idx = url_cell.row
                
                if key and key in prog_data_map:
                    data = prog_data_map[key]
                    
                    # 写入预设问题
                    ws_out.cell(row=current_row_idx, column=o_col_q1, value=data['questions'][0])
                    ws_out.cell(row=current_row_idx, column=o_col_q2, value=data['questions'][1])
                    ws_out.cell(row=current_row_idx, column=o_col_q3, value=data['questions'][2])
                    
                    # 写入机械标题
                    ws_out.cell(row=current_row_idx, column=mech_title_col_idx, value=data['section_title'])
                    
                    # 修正视频URL：使用程序表中规范的URL覆盖人工表中的URL
                    # 注意：o_idx_url_col_num 是列号（1-based），直接使用
                    raw_prog_url = data.get('raw_url')
                    if raw_prog_url:
                        ws_out.cell(row=current_row_idx, column=o_idx_url_col_num, value=raw_prog_url)
                    
                    matched_count += 1
                else:
                    pass
                    # print(f"未匹配: {url} (Key: {key})")

            print(f"处理行数: {processed_count}")
            print(f"成功匹配并更新: {matched_count} 行")
            
            if matched_count == 0:
                print("\n❌ 警告: 匹配数量为0！可能原因：")
                print("1. 人工表和程序表的URL格式差异过大")
                print("2. 列索引识别错误")
                print("\n--- 调试信息 ---")
                print("人工表前3个URL Key:")
                count = 0
                for row in ws_out.iter_rows(min_row=2, max_row=4):
                     if len(row) >= o_idx_url_col_num:
                         u = row[o_idx_url_col_num-1].value
                         print(f"  URL: {u} -> Key: {get_video_key(u)}")
                print("\n程序表前3个Key:")
                for k in list(prog_data_map.keys())[:3]:
                    print(f"  Key: {k} (Source: {prog_data_map[k]['raw_url']})")
            
        else:
            print("❌ 错误: 未找到 chapters_sections 分表")

        # ==========================================
        # 处理 exercises 分表
        # ==========================================
        if 'exercises' in wb_prog.sheetnames:
            print("\n正在处理 exercises 分表...")
            ws_prog_ex = wb_prog['exercises']
            
            if 'exercises' in wb_out.sheetnames:
                del wb_out['exercises']
            
            ws_out_ex = wb_out.create_sheet('exercises')
            
            # 复制所有内容
            for row in ws_prog_ex.iter_rows(values_only=True):
                ws_out_ex.append(row)
            print("✅ exercises 分表已替换")
        else:
            print("⚠️ 警告: 程序Excel中未找到 exercises 分表")

        # ==========================================
        # 保存结果
        # ==========================================
        print(f"\n正在保存文件...")
        wb_out.save(output_file)
        print(f"🎉 合并完成！请查看: {output_file}")

    except Exception as e:
        print(f"❌ 发生未预期的错误: {e}")
        import traceback
        traceback.print_exc()

def main():
    parser = argparse.ArgumentParser(description="合并人工编辑的Excel和程序生成的Excel")
    parser.add_argument("manual_excel", help="人工编辑的Excel文件路径")
    parser.add_argument("program_excel", help="程序生成的Excel文件路径")
    parser.add_argument("-o", "--output", help="输出文件路径 (默认为 merged_output.xlsx)")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.manual_excel):
        print(f"错误: 人工Excel文件不存在: {args.manual_excel}")
        return
    if not os.path.exists(args.program_excel):
        print(f"错误: 程序Excel文件不存在: {args.program_excel}")
        return
        
    merge_excel_files(args.manual_excel, args.program_excel, args.output)

if __name__ == "__main__":
    main()
