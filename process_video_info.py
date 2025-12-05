import json
import shutil
import openpyxl
import math
import re
import os
import threading

# 确保这个锁在文件顶层被定义，它就是唯一的、共享的实例
excel_file_lock = threading.Lock()

def sanitize_filename(filename):
    """
    移除或替换 Windows 文件名中的非法字符。
    非法字符包括: \ / : * ? " < > |
    """
    # 将所有非法字符替换为下划线 '_'
    return re.sub(r'[\\/:*?"<>| ]', '_', filename)

def process_video_to_excel_flash(json_file_path, template_excel_path, video_index, part_title=None, page_num=None):
    """
    批量处理时，将所有信息追加到以父目录命名的 Excel 文件中。
    Args:
        json_file_path: json文件路径
        template_excel_path: 模板excel路径
        video_index: 视频索引（虽然函数内似乎没用到这个参数作为行号，而是查找空行）
        part_title: 分P标题（可选），如果提供则使用此标题代替主标题
        page_num: 分P序号（可选），如果提供且大于1，则拼接到URL后面
    """
    # 1. 读取并解析 JSON 文件
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            video_info = json.load(f)
        print(f"成功读取 JSON 文件: {json_file_path}")
    except FileNotFoundError:
        print(f"错误: JSON 文件未找到 '{json_file_path}'")
        return
    except json.JSONDecodeError:
        print(f"错误: JSON 文件格式不正确 '{json_file_path}'")
        return

    # 获取父目录名作为 Excel 文件名
    parent_dir = os.path.basename(os.path.dirname(json_file_path))
    excel_filename = os.path.join(os.path.dirname(json_file_path), f"{sanitize_filename(parent_dir)}.xlsx")

    # 如果 Excel 文件不存在，则复制模板
    if not os.path.exists(excel_filename):
        shutil.copy(template_excel_path, excel_filename)
        print(f"模板文件已复制并重命名为: {excel_filename}")
    # 打开 Excel 文件
    try:
        with excel_file_lock:
            workbook = openpyxl.load_workbook(excel_filename)
            sheet_chapters = workbook['chapters_sections'] if 'chapters_sections' in workbook.sheetnames else None
            if not sheet_chapters:
                print("警告: 未找到 'chapters_sections' 工作表。")
                return

            # 找到第一个空行（避免使用 video_index 作为行号，因为在多线程环境中
            # video_index 可能是线程名如 'Worker-1'，仅包含少量不同数字，会导致重复写入相同行）
            # 从第2行开始（假设第1行为表头）向下查找第一个空白的序号列
            current_row = 2
            while True:
                cell_val = sheet_chapters.cell(row=current_row, column=1).value
                if cell_val in (None, ''):
                    break
                current_row += 1

            # 处理分P信息
            bvid = video_info.get('bvid')
            duration_sec = video_info.get('duration', 0)
            duration_min = math.ceil(duration_sec / 60) if duration_sec > 0 else 0
            
            # 确定使用的标题：如果提供了分P标题则使用，否则使用视频主标题
            title_to_use = part_title if part_title else video_info.get('title', '')
            
            # 构建URL
            video_url = f"https://www.bilibili.com/video/{bvid}"
            if page_num and int(page_num) > 1:
                video_url += f"/?p={page_num}"
            
            sheet_chapters.cell(row=current_row, column=1, value=current_row - 1)
            sheet_chapters.cell(row=current_row, column=4, value=video_url)
            sheet_chapters.cell(row=current_row, column=5, value=sanitize_filename(title_to_use))
            sheet_chapters.cell(row=current_row, column=6, value=current_row - 1)
            sheet_chapters.cell(row=current_row, column=7, value=duration_min)

            workbook.save(excel_filename)
            print(f"数据已追加并保存到文件: {excel_filename}")

    except Exception as e:
        print(f"处理 Excel 文件时发生错误: {e}")

def process_video_to_excel_final(json_file_path, template_excel_path):
    """
    读取B站视频信息json，净化标题作为文件名，并将原始数据填充到Excel模板中，同时保持模板格式。

    Args:
        json_file_path (str): video_info.json 文件的路径。
        template_excel_path (str): 模板.xlsx 文件的路径。
    """
    # 1. 读取并解析 JSON 文件
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            video_info = json.load(f)
        print(f"成功读取 JSON 文件: {json_file_path}")
    except FileNotFoundError:
        print(f"错误: JSON 文件未找到 '{json_file_path}'")
        return
    except json.JSONDecodeError:
        print(f"错误: JSON 文件格式不正确 '{json_file_path}'")
        return

    # 2. 获取原始标题，video_info.json相同路径下创建 Excel 文件
    original_title = sanitize_filename(video_info.get('title', 'Untitled Course'))
    # 处理带前缀的video_info.json文件名（例如：标题_video_info.json）
    safe_excel_filename = json_file_path.replace('_video_info.json', '.xlsx')
    # 如果没有前缀（旧格式），则使用原来的逻辑
    if safe_excel_filename == json_file_path:
        safe_excel_filename = json_file_path.replace('video_info.json', f'{original_title}.xlsx')

    # 3. 复制并重命名模板 Excel 文件
    try:
        shutil.copy(template_excel_path, safe_excel_filename)
        print(f"模板文件已复制并重命名为: {safe_excel_filename}")
    except FileNotFoundError:
        print(f"错误: 模板 Excel 文件未找到 '{template_excel_path}'")
        return
    except Exception as e:
        print(f"复制文件时出错: {e}")
        return

    try:
        # 4. 使用 openpyxl 打开新的 Excel 文件并填入数据
        workbook = openpyxl.load_workbook(safe_excel_filename)

        # 5. 填充 'course' 工作表 (使用原始数据)
        if 'course' in workbook.sheetnames:
            sheet_course = workbook['course']
            sheet_course.cell(row=2, column=1, value=1)  # 序号
            sheet_course.cell(row=2, column=2, value=original_title)  # 课程名称 (原始标题)
            sheet_course.cell(row=2, column=3, value=video_info.get('pic', ''))  # 课程图标URL
            sheet_course.cell(row=2, column=4, value=video_info.get('desc', ''))  # 课程描述
            print("'course' 工作表数据填充完毕。")
        else:
            print("警告: 在模板文件中未找到名为 'course' 的工作表。")

        # 处理合集/系列视频的分P信息
        if 'ugc_season' in video_info and video_info['ugc_season'].get('sections'):
            pages = []
            # 这是一个合集/系列。遍历所有 sections 中的 episodes
            for section in video_info['ugc_season']['sections']:
                if section.get('episodes'):
                    # 将每个 episode 转换为与原 pages 结构类似的字典，方便后续逻辑复用
                    for episode in section['episodes']:
                        # 提取关键信息，bvid 是下载字幕时可能需要传递的参数
                        cid = episode.get('cid')
                        page_title = episode.get('title') or episode.get('page', {}).get('part', '未知剧集')
                        duration_sec = episode.get('pages', [])[0].get('duration', 0)
                        duration_min = math.ceil(duration_sec / 60) if duration_sec > 0 else 0
                        bvid_ep = episode.get('bvid')

                        if cid:
                            pages.append({
                                'cid': cid,
                                'part': page_title,
                                'bvid': bvid_ep,  # 存储bvid，以便在循环中使用
                                'duration': duration_min
                            })
        else:
            # 否则，使用顶层的 'pages' 字段作为分P列表（如果是普通视频）
            pages = video_info.get('pages', [])
            if not pages:
                print("错误: 未找到视频分P信息")
                return

        # 6. 填充 'chapters_sections' 工作表 (使用原始数据)
        if 'chapters_sections' in workbook.sheetnames:
            sheet_chapters = workbook['chapters_sections']
            bvid = video_info.get('bvid')
            if video_info.get('ugc_season'):
                for index, episode in enumerate(pages):
                    current_row = index + 2
                    sheet_chapters.cell(row=current_row, column=1, value=index + 1)
                    sheet_chapters.cell(row=current_row, column=4,
                                        value=f"https://www.bilibili.com/video/{episode.get('bvid')}")
                    sheet_chapters.cell(row=current_row, column=5,value=sanitize_filename(episode.get('part', '')))
                    sheet_chapters.cell(row=current_row, column=6, value=index + 1)
                    sheet_chapters.cell(row=current_row, column=7, value=episode.get('duration'))
            else:
                for index, page in enumerate(pages):
                    current_row = index + 2
                    duration_sec = page.get('duration', 0)
                    duration_min = math.ceil(duration_sec / 60) if duration_sec > 0 else 0

                    sheet_chapters.cell(row=current_row, column=1, value=index + 1)
                    sheet_chapters.cell(row=current_row, column=4,
                                        value=f"https://www.bilibili.com/video/{bvid}/?p={page.get('page')}")
                    sheet_chapters.cell(row=current_row, column=5, value=sanitize_filename(page.get('part', '')))  # 节标题 (原始数据)
                    sheet_chapters.cell(row=current_row, column=6, value=page.get('page', index + 1))
                    sheet_chapters.cell(row=current_row, column=7, value=duration_min)
                print(f"'chapters_sections' 工作表数据填充完毕，共 {len(pages)} 条记录。")
        else:
            print("警告: 在模板文件中未找到名为 'chapters_sections' 的工作表。")

        # 7. 保存修改后的 Excel 文件
        workbook.save(safe_excel_filename)
        print(f"数据已成功写入并保存到文件: {safe_excel_filename}")

    except Exception as e:
        print(f"处理 Excel 文件时发生错误: {e}")


if __name__ == '__main__':
    # 定义文件路径
    json_file = 'video_info.json'
    template_file = '模板.xlsx'

    # 执行主函数
    process_video_to_excel_final(json_file, template_file)