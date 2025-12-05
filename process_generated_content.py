import os
import json
from openpyxl import load_workbook
import argparse

def load_json(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)

def fill_questions_sheet(ws_cs, section_title, questions):
    """
    填充 chapters_sections 表中的 预设问题 列（同之前逻辑，不变）
    """
    col_title = "节标题"
    preset_cols = ["预设问题1", "预设问题2", "预设问题3"]
    header = {cell.value: cell.column for cell in ws_cs[1]}
    if col_title not in header:
        raise ValueError(f"在 chapters_sections 表中没有列 {col_title}")
    for pc in preset_cols:
        if pc not in header:
            raise ValueError(f"在 chapters_sections 表中没有列 {pc}")
    title_col_idx = header[col_title]
    preset_col_idxs = [ header[pc] for pc in preset_cols ]

    for row in ws_cs.iter_rows(min_row=2):
        cell = row[title_col_idx - 1]
        if cell.value == section_title:
            for q in questions:
                qid = q.get("id")
                qtext = q.get("question")
                if isinstance(qid, int) and 1 <= qid <= len(preset_col_idxs):
                    col_idx = preset_col_idxs[qid - 1]
                    row[col_idx - 1].value = qtext
            break

def fill_exercises_sheet(ws_ex, section_title, exercises, start_serial):
    """
    填充 exercises 表，返回下一个可用的 序号（start_serial 是传入的起始序号）。
    此函数会追加行，每个题目一行。

    参数说明：
    - ws_ex: exercises sheet worksheet
    - section_title: 当前处理的节标题
    - exercises: JSON 解析出来的 dict，可能含多个题型
    - start_serial: 传入这个 section 开始时的序号编号（int），用于给这一节的题目编号
    返回值是填完后返回的下一个可用序号（用于下一节接着写序号连续）
    """
    # Excel 中列名
    col_serial = "序号"
    col_section = "节标题"
    col_qbody = "习题正文"
    col_type = "题型(单选/多选/简答)"
    col_score = "分值"
    col_answer = "正确答案"
    # 选项列 A〜G
    opt_cols = ["选项A", "选项B", "选项C", "选项D", "选项E", "选项F", "选项G"]

    header = {cell.value: cell.column for cell in ws_ex[1]}
    # 必要列检查
    needed = [col_serial, col_section, col_qbody, col_type, col_score, col_answer] + opt_cols
    for c in needed:
        if c not in header:
            raise ValueError(f"在 exercises 表中没有列 {c}")

    max_row = ws_ex.max_row
    serial = start_serial

    # 处理 multiple_choice 题型（可视为单选或多选，根据 JSON 结构看是“多选题 / 多选”或“单选题 / 单选”）
    for mc in exercises.get("multiple_choice", []):
        qtext = mc.get("question")
        opts = mc.get("options", {})  # dict A/B/C/...
        correct = mc.get("correct_answer")
        # 题型在 Excel 中填 “多选” 或 “单选” —— 如果 JSON 里指明是多选（例如有多个正确答案）可判定
        # 此处假设 mc 是“多选题”或“单选题”，用 “多选” 或 “单选”
        qtype = "多选"  # 默认标为多选
        # 如果正确答案是单个字符（如 "A" / "B"），也可以把它归为单选
        if isinstance(correct, str) and len(correct) == 1:
            qtype = "单选"
        # 分值统一设为 5
        score = 5

        # 写入一行
        new_row = max_row + 1
        ws_ex.cell(row=new_row, column=header[col_serial], value=serial)
        ws_ex.cell(row=new_row, column=header[col_section], value=section_title)
        ws_ex.cell(row=new_row, column=header[col_qbody], value=qtext)
        ws_ex.cell(row=new_row, column=header[col_type], value=qtype)
        ws_ex.cell(row=new_row, column=header[col_score], value=score)
        ws_ex.cell(row=new_row, column=header[col_answer], value=correct)
        # 写选项
        for i, opt_col in enumerate(opt_cols):
            key = chr(ord('A') + i)  # 'A', 'B', 'C', ...
            if key in opts:
                ws_ex.cell(row=new_row, column=header[opt_col], value=opts.get(key))
            else:
                # 如果这一列没有选项，就不写（保持空白）
                pass
        max_row += 1
        serial += 1

    # 处理简答题 short_answer
    for sa in exercises.get("short_answer", []):
        qtext = sa.get("question")
        # JSON 里给的参考答案或答案要点
        # 假设我们把 reference_answer 或 join(answer_points) 作为正确答案填入 Excel
        reference = sa.get("reference_answer")
        if not reference:
            # 如果没有 reference_answer，也可拼接 answer_points
            ap = sa.get("answer_points", [])
            reference = "\n".join(ap)
        correct = reference
        qtype = "简答"
        # 分值统一设为 15
        score = 15

        new_row = max_row + 1
        ws_ex.cell(row=new_row, column=header[col_serial], value=serial)
        ws_ex.cell(row=new_row, column=header[col_section], value=section_title)
        ws_ex.cell(row=new_row, column=header[col_qbody], value=qtext)
        ws_ex.cell(row=new_row, column=header[col_type], value=qtype)
        ws_ex.cell(row=new_row, column=header[col_score], value=score)
        ws_ex.cell(row=new_row, column=header[col_answer], value=correct)
        # 选项列保持空白（简答题无选项 A〜G）

        max_row += 1
        serial += 1

    # 如有其他题型（比如判断、填空等）也可以在这里类似扩展处理

    return serial  # 返回写完后的下一个可用序号

def find_matching_file(directory, suffix):
    """
    在指定目录中查找以 suffix 结尾的文件
    """
    if not os.path.exists(directory):
        return None
    
    for filename in os.listdir(directory):
        if filename.endswith(suffix):
            return os.path.join(directory, filename)
    return None

def save_data_to_excel(excel_filename):
    relative_directory = os.path.dirname(excel_filename)
    wb = load_workbook(excel_filename)
    ws_cs = wb["chapters_sections"]
    ws_ex = wb["exercises"]
    ws_ex.delete_rows(2, ws_ex.max_row - 1)  # 清空 exercises 表中除表头外的所有行
    # 先处理 chapters_sections sheet，把预设问题填进去
    header_cs = {cell.value: cell.column for cell in ws_cs[1]}
    if "节标题" not in header_cs:
        raise ValueError("在 chapters_sections 表中找不到 “节标题” 列")
    title_col_idx = header_cs["节标题"]

    # 为 exercises 表准备一个全局序号计数器，从 1 开始
    next_serial = 1

    for row in ws_cs.iter_rows(min_row=2):
        cell = row[title_col_idx - 1]
        section_title = cell.value
        if not section_title:
            continue

        # 尝试直接查找
        q_fn = f"{relative_directory}/{section_title}_questions.json"
        ex_fn = f"{relative_directory}/{section_title}_exercises.json"
        
        # 如果直接查找失败，尝试通过后缀查找（处理分P文件名带前缀的情况）
        if not os.path.exists(q_fn):
            found_q = find_matching_file(relative_directory, f"{section_title}_questions.json")
            if found_q:
                q_fn = found_q
                
        if not os.path.exists(ex_fn):
            found_ex = find_matching_file(relative_directory, f"{section_title}_exercises.json")
            if found_ex:
                ex_fn = found_ex

        # 填 questions 部分
        if os.path.exists(q_fn):
            j = load_json(q_fn)
            questions = j.get("questions", [])
            fill_questions_sheet(ws_cs, section_title, questions)
        else:
            print(f"未找到文件：{q_fn}")

        # 填 exercises 部分
        if os.path.exists(ex_fn):
            j2 = load_json(ex_fn)
            next_serial = fill_exercises_sheet(ws_ex, section_title, j2, next_serial)
        else:
            print(f"未找到文件：{ex_fn}")

    # 保存
    wb.save(excel_filename)
    print("已完成写入并保存 Excel 文件。")

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='字幕总结工具 - 使用大模型分析字幕并生成要点总结',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
    使用示例:
      # 基本使用 方法
      python process_generated_content.py subtitles/填课程名称/填表名称.xlsx
            """
    )

    parser.add_argument('excel_file', nargs='?', help='excel文件路径')

    args = parser.parse_args()

    # 检查是否提供了excel文件
    if not args.excel_file:
        parser.print_help()
        print("\n错误：请按照提示提供excel文件")
        return

    excel_filename = args.excel_file
    # 检查excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误：excel文件不存在: {args.excel_file}")
        return

    save_data_to_excel(excel_filename)


if __name__ == "__main__":
    main()
